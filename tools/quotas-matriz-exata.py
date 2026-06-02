"""
Cria recibos virtuais aplicando a regra:
- Pagamento de N×quota_mensal cobre N meses consecutivos
- Começa no primeiro mês não coberto do ano e estende sequencialmente
- A vista de Quotas mostra a quota_mensal em cada mês COBERTO (não o valor agregado do pagamento)
"""
import openpyxl, json
from datetime import datetime
from collections import Counter

EXCEL = "/mnt/user-data/uploads/Contas_Condominio_2023_2026_-_actualizado_a_25-05-2026.xlsx"
SEED = "/home/claude/GestaoCondominioAR24/data/seed-historico.json"

COL_TO_TENANT = {1:'cond_02',2:'cond_01',3:'cond_04',4:'cond_03',5:'cond_06',
                 6:'cond_05',7:'cond_08',8:'cond_07',9:'cond_10',10:'cond_09'}

MES_PT = ['Janeiro','Fevereiro','Março','Abril','Maio','Junho',
          'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']

wb = openpyxl.load_workbook(EXCEL, data_only=True)

# Ler quotas mensais por ano (L5 do Excel principal)
quota_mensal = {}  # ano → tenant → quota_cent
for ano in ['2024','2025','2026']:
    ws = wb[f'Quotas {ano}']
    row = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))[0]
    quota_mensal[ano] = {}
    for ci, tid in COL_TO_TENANT.items():
        v = row[ci]
        if v and isinstance(v, (int, float)):
            quota_mensal[ano][tid] = int(round(v * 100))

print("Quotas mensais por ano:")
for ano in ['2024','2025','2026']:
    print(f"  {ano}: {quota_mensal[ano]}")

# Para cada (tenant, ano), iterar pelos meses cronologicamente
# Cada pagamento (célula > 0) cobre N meses = valor / quota_mensal
celulas_normalizadas = []  # (ano, mes_coberto, tid, valor_cent_quota_mensal)
saldos_excesso = []  # registar valor que excede 12 meses (raro)

for ano in ['2024','2025','2026']:
    ws = wb[f'Quotas {ano}']
    for tid in COL_TO_TENANT.values():
        qm = quota_mensal[ano].get(tid)
        if not qm: continue
        # Ler valores mensais
        pagamentos_mes = {}  # mes → valor_cent
        for mes in range(1, 13):
            row = list(ws.iter_rows(min_row=mes+9, max_row=mes+9, values_only=True))[0]
            col_idx = [k for k,v in COL_TO_TENANT.items() if v==tid][0]
            v = row[col_idx]
            if v and isinstance(v,(int,float)) and v > 0:
                pagamentos_mes[mes] = int(round(v*100))

        # Distribuir: cronológico, cada pagamento cobre N meses sequenciais
        proximo_mes_a_cobrir = 1
        for mes_pgto in sorted(pagamentos_mes.keys()):
            valor_cent = pagamentos_mes[mes_pgto]
            # N meses = valor / quota_mensal (arredondado para mais próximo)
            n_meses = round(valor_cent / qm)
            if n_meses == 0: n_meses = 1  # mesmo se < quota inteira, cobre 1 mês
            for i in range(n_meses):
                mes_coberto = proximo_mes_a_cobrir + i
                if mes_coberto > 12:
                    saldos_excesso.append((ano, tid, mes_coberto, qm))
                    continue
                celulas_normalizadas.append((ano, mes_coberto, tid, qm))
            proximo_mes_a_cobrir += n_meses

print(f"\nCélulas cobertas (após normalização): {len(celulas_normalizadas)}")
if saldos_excesso:
    print(f"⚠ Saldos a favor (meses > 12): {len(saldos_excesso)}")

# Carregar JSON
with open(SEED) as f:
    data = json.load(f)

# Apagar quotas antigas, manter prestação/outros
nao_quota = [r for r in data['receipts'] if r['tipo'] != 'quota']
print(f"Recibos não-quota mantidos: {len(nao_quota)}")

tenants_map = {t['id']: t for t in data['tenants']}

# Criar virtuais
virtuais = []
seq_por_ano = Counter()
for ano, mes, tid, valor_cent in sorted(celulas_normalizadas, key=lambda c:(c[0],c[1],c[2])):
    seq_por_ano[ano] += 1
    seq = seq_por_ano[ano]
    tenant = tenants_map[tid]
    virtuais.append({
        'id': f'rcb_h_{ano}_{seq:03d}',
        'recibo_numero': f'H{seq:03d}/ADM{ano}',
        'ano': ano,
        'tenantId': tid,
        'tenantName': tenant['name'],
        'fraction': tenant.get('fraction', ''),
        'data': f"{ano}-{mes:02d}-15",
        'valor_centimos': valor_cent,
        'mesReferencia': [f"{ano}-{mes:02d}"],
        'descricao': f'Quota mensal {MES_PT[mes-1]} {ano}',
        'tipo': 'quota',
        'criadoEm': datetime(int(ano), mes, 15).timestamp() * 1000,
        'historico': True,
    })

# Combinar
data['receipts'] = sorted(virtuais + nao_quota, key=lambda r:(r['ano'], r['data']))
print(f"Recibos virtuais criados: {len(virtuais)}")

# nextNumberByYear baseado nos RCB reais
import re
max_rcb = {}
for r in nao_quota:
    m = re.match(r'RCB\s+(\d+)/ADM(\d{4})', r['recibo_numero'])
    if m: 
        n,a = int(m.group(1)),m.group(2)
        max_rcb[a] = max(max_rcb.get(a,0), n)
data['meta']['config']['nextNumberByYear'] = {a:n+1 for a,n in max_rcb.items()}
print(f"\nnextNumberByYear: {data['meta']['config']['nextNumberByYear']}")

# Salvar
with open(SEED, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

# Mostrar tabela 2026 resultante
print("\n=== Tabela 2026 RESULTANTE ===")
mat = {tid: {m:0 for m in range(1,13)} for tid in COL_TO_TENANT.values()}
for c in celulas_normalizadas:
    if c[0]=='2026': mat[c[2]][c[1]] += c[3]

NOMES = {'cond_01':'João','cond_02':'Filipe','cond_03':'Leonel','cond_04':'Sílvia',
         'cond_05':'Ricardo','cond_06':'António','cond_07':'Nuno','cond_08':'Lurdes',
         'cond_09':'Monteiro','cond_10':'Vitor/Ana'}
print(f"{'Tenant':<10}" + "".join(f"{m:>5}" for m in ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']) + "  Total")
for tid, nome in NOMES.items():
    row = mat[tid]
    total = sum(row.values())
    print(f"  {nome:<8}" + "".join(f"{(row[m]/100):>5.0f}" if row[m] else "    ·" for m in range(1,13)) + f"  {total/100:>6.0f}")
