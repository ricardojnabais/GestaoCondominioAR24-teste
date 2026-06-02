"""
Gera seed-historico.json a partir do backup Excel.
Estrutura: tenants, rubricas, receipts, pagamentosDespesa, planos, prestacoes, orcamentos, meta.
"""
import openpyxl
import json
from datetime import datetime
import os, sys

EXCEL = "/mnt/user-data/uploads/Contas_Condominio_2023_2026_-_actualizado_a_25-05-2026.xlsx"
OUT = "/home/claude/GestaoCondominioAR24/data/seed-historico.json"
os.makedirs(os.path.dirname(OUT), exist_ok=True)

wb = openpyxl.load_workbook(EXCEL, data_only=True)

# ──────────────────────────────────────────────────────────
# MAPEAMENTOS
# ──────────────────────────────────────────────────────────

# Coluna no Excel (0-based) → tenantId
# Excel: posição 0 = label "Fracções"
#        posição 1 = R/C.Dto, 2 = R/C.Esq, 3 = 1º Dto, ...
COL_TO_TENANT = {
    1: 'cond_02',   # R/C Dto - Filipe Solha (91‰) - ADMIN
    2: 'cond_01',   # R/C Esq - João Vaz (79‰)
    3: 'cond_04',   # 1º Dto - Sílvia Gonçalves (87‰) ⚠ em atraso
    4: 'cond_03',   # 1º Esq - Leonel Venâncio (119‰)
    5: 'cond_06',   # 2º Dto - António Figueiredo (88‰)
    6: 'cond_05',   # 2º Esq - Ricardo Cordeiro (121‰)
    7: 'cond_08',   # 3º Dto - Lurdes Serafim (88‰)
    8: 'cond_07',   # 3º Esq - Nuno Silva (115‰)
    9: 'cond_10',   # 4º Dto - Vitor Barata (87‰)
    10: 'cond_09',  # 4º Esq - José Carlos Monteiro (125‰)
}

# Tenants completos (com rentByYear de cada ano)
TENANTS = [
    {'id': 'cond_01', 'name': 'João Vaz',           'fraction': 'R/C Esquerdo', 'permilage': 79,
     'nif': '129465380', 'email': 'joaovaz@example.com',          'isAdmin': False,
     'rentByYear': {'2021': 2800, '2022': 2800, '2023': 3200, '2024': 3200, '2025': 3200, '2026': 3200}},
    {'id': 'cond_02', 'name': 'Filipe Solha',       'fraction': 'R/C Direito',  'permilage': 91,
     'nif': '219481342', 'email': 'filipesolha@gmail.com',         'isAdmin': True,
     'rentByYear': {'2021': 3200, '2022': 3200, '2023': 3700, '2024': 3700, '2025': 3700, '2026': 3700}},
    {'id': 'cond_03', 'name': 'Leonel Venâncio',    'fraction': '1.º Esquerdo', 'permilage': 119,
     'nif': '209959746', 'email': 'leonelvenancio@example.com',    'isAdmin': False,
     'rentByYear': {'2021': 4200, '2022': 4200, '2023': 4800, '2024': 4800, '2025': 4800, '2026': 4800}},
    {'id': 'cond_04', 'name': 'Sílvia Gonçalves',   'fraction': '1.º Direito',  'permilage': 87,
     'nif': '195084381', 'email': 'silviagoncalves@example.com',   'isAdmin': False,
     'rentByYear': {'2021': 3100, '2022': 3100, '2023': 3600, '2024': 3600, '2025': 3600, '2026': 3600}},
    {'id': 'cond_05', 'name': 'Ricardo Cordeiro',   'fraction': '2.º Esquerdo', 'permilage': 121,
     'nif': '214490041', 'email': 'ricardojnabais@gmail.com',      'isAdmin': True,
     'rentByYear': {'2021': 4300, '2022': 4300, '2023': 4900, '2024': 4900, '2025': 4900, '2026': 4900}},
    {'id': 'cond_06', 'name': 'António Figueiredo', 'fraction': '2.º Direito',  'permilage': 88,
     'nif': '101744137', 'email': 'antoniofigueiredo@example.com', 'isAdmin': False,
     'rentByYear': {'2021': 3100, '2022': 3100, '2023': 3600, '2024': 3600, '2025': 3600, '2026': 3600}},
    {'id': 'cond_07', 'name': 'Nuno Silva',         'fraction': '3.º Esquerdo', 'permilage': 115,
     'nif': '195611004', 'email': 'nunosilva@example.com',         'isAdmin': False,
     'rentByYear': {'2021': 4100, '2022': 4100, '2023': 4700, '2024': 4700, '2025': 4700, '2026': 4700}},
    {'id': 'cond_08', 'name': 'Lurdes Serafim',     'fraction': '3.º Direito',  'permilage': 88,
     'nif': '127143980', 'email': 'lurdesserafim@example.com',     'isAdmin': False,
     'rentByYear': {'2021': 3100, '2022': 3100, '2023': 3600, '2024': 3600, '2025': 3600, '2026': 3600}},
    {'id': 'cond_09', 'name': 'José Carlos Monteiro','fraction': '4.º Esquerdo', 'permilage': 125,
     'nif': '182258637', 'email': 'jcmonteiro@example.com',        'isAdmin': False,
     'rentByYear': {'2021': 4400, '2022': 4400, '2023': 5100, '2024': 5100, '2025': 5100, '2026': 5100}},
    {'id': 'cond_10', 'name': 'Vitor Barata',       'fraction': '4.º Direito',  'permilage': 87,
     'nif': '178132730', 'email': 'vitorbarata@example.com',       'isAdmin': False,
     'rentByYear': {'2021': 3100, '2022': 3100, '2023': 3600, '2024': 3600, '2025': 3600, '2026': 3600}},
]

# ──────────────────────────────────────────────────────────
# RÚBRICAS
# ──────────────────────────────────────────────────────────
RUBRICAS = [
    {'id': 'rub_edp',          'nome': 'EDP / Electricidade',  'cor': '#F59E0B', 'ativa': True, 'ordem': 1},
    {'id': 'rub_telefone',     'nome': 'Telefone',             'cor': '#06B6D4', 'ativa': False, 'ordem': 2,
     'descricao': 'Linha telefónica do prédio (terminou em 2022)'},
    {'id': 'rub_schindler',    'nome': 'Schindler / Elevador', 'cor': '#1E54C7', 'ativa': True, 'ordem': 3},
    {'id': 'rub_agua',         'nome': 'Água',                 'cor': '#3B82F6', 'ativa': True, 'ordem': 4},
    {'id': 'rub_limpeza',      'nome': 'Limpeza',              'cor': '#10B981', 'ativa': True, 'ordem': 5},
    {'id': 'rub_banc',         'nome': 'Despesas Bancárias',   'cor': '#6B7280', 'ativa': True, 'ordem': 6},
    {'id': 'rub_allianz',      'nome': 'Allianz / Seguros',    'cor': '#7C3AED', 'ativa': True, 'ordem': 7},
    {'id': 'rub_obras',        'nome': 'Obras / Intervenções', 'cor': '#DC2626', 'ativa': True, 'ordem': 8},
    {'id': 'rub_plano_schindler', 'nome': 'Plano Schindler', 'cor': '#0891B2', 'ativa': True, 'ordem': 9,
     'fornecedorDefault': 'Schindler', 'valorDefault': 42796,
     'descricao': 'Plano de pagamento faseado para reparação do elevador',
     'metodoPagamentoDefault': 'transferencia'},
    {'id': 'rub_outras',       'nome': 'Outras',               'cor': '#9CA3AF', 'ativa': True, 'ordem': 10},
]

# ──────────────────────────────────────────────────────────
# QUOTAS HISTÓRICAS · gera recibos por célula
# ──────────────────────────────────────────────────────────

MES_NAMES = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
             'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']

receipts = []
receipt_counter_by_year = {}

def next_num(year):
    receipt_counter_by_year[year] = receipt_counter_by_year.get(year, 0) + 1
    return f"H{receipt_counter_by_year[year]:03d}/{year}"

def parse_quotas_sheet(sheet_name, year):
    """Lê uma folha de Quotas YYYY e gera recibos.
    Linhas: 11-22 = Jan-Dez Receitas (em 2024-2026 começa na linha 10)."""
    ws = wb[sheet_name]
    # Encontrar onde começa "Receitas" (header linha) e onde começa Janeiro
    rows = list(ws.iter_rows(values_only=True))
    start_row = None
    for i, row in enumerate(rows):
        if row and len(row) > 0 and row[0] and str(row[0]).strip().startswith('Receitas'):
            start_row = i + 1  # Linha seguinte = Janeiro
            break
    if start_row is None:
        # Fallback: padrão
        start_row = 10 if year >= 2024 else 10  # 0-indexed
    # 12 meses
    for mes_idx in range(12):
        row = rows[start_row + mes_idx] if start_row + mes_idx < len(rows) else None
        if not row:
            continue
        # Coluna 0 deve ser o nome do mês (ou data)
        for col_idx, tenant_id in COL_TO_TENANT.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is None or val == '---' or val == '':
                continue
            try:
                val_eur = float(val)
            except (TypeError, ValueError):
                continue
            if val_eur <= 0:
                continue
            mes_num = mes_idx + 1
            valor_cent = int(round(val_eur * 100))
            data = f"{year}-{mes_num:02d}-15"  # convenção meio do mês
            tenant = next(t for t in TENANTS if t['id'] == tenant_id)
            quota_mensal = tenant['rentByYear'][str(year)]
            # Determinar coverage: assumimos cobre meses começando do próprio mês
            coverage = []
            restante = valor_cent
            ano_cov = year
            mes_cov = mes_num
            while restante >= quota_mensal:
                coverage.append({'year': ano_cov, 'month': mes_cov, 'valor_centimos': quota_mensal})
                restante -= quota_mensal
                mes_cov += 1
                if mes_cov > 12:
                    mes_cov = 1
                    ano_cov += 1
            if restante > 0:
                coverage.append({'year': ano_cov, 'month': mes_cov, 'valor_centimos': restante})
            receipts.append({
                'id': f'rcp_h_{year}_{mes_num:02d}_{tenant_id}',
                'recibo_numero': next_num(year),
                'ano': str(year),
                'tenantId': tenant_id,
                'data': data,
                'valor_centimos': valor_cent,
                'mesReferencia': [f"{c['year']}-{c['month']:02d}" for c in coverage],
                'descricao': f'Quota {MES_NAMES[mes_idx]} {year} (histórico)',
                'tipo': 'quota',
                'historico': True,
                'criadoEm': datetime(year, mes_num, 15).timestamp() * 1000,
            })

for year, sheet_name in [(2021,'Quotas 2021'),(2022,'Quotas 2022'),(2023,'Quotas 2023'),
                          (2024,'Quotas 2024'),(2025,'Quotas 2025'),(2026,'Quotas 2026')]:
    parse_quotas_sheet(sheet_name, year)

print(f"Recibos gerados: {len(receipts)}")

# ──────────────────────────────────────────────────────────
# DESPESAS HISTÓRICAS · cada célula vira pagamento de despesa
# ──────────────────────────────────────────────────────────

pagamentosDespesa = []
despesa_counter = 0

def add_despesa(year, mes, val_eur, rubricaId, descricao):
    global despesa_counter
    if val_eur is None or val_eur == 0:
        return
    try:
        val_eur = float(val_eur)
    except (TypeError, ValueError):
        return
    if val_eur <= 0:
        return
    despesa_counter += 1
    pagamentosDespesa.append({
        'id': f'pd_h_{year}_{mes:02d}_{despesa_counter}',
        'rubricaId': rubricaId,
        'valor_centimos': int(round(val_eur * 100)),
        'data': f"{year}-{mes:02d}-15",
        'descricao': descricao,
        'metodoPagamento': 'transferencia',
        'fornecedor': '',
        'criadoEm': datetime(year, mes, 15).timestamp() * 1000,
        'historico': True,
    })

def parse_despesas(sheet_name, year):
    """Lê folha Despesas YYYY. Encontra mapa rúbricas → coluna."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    # Encontrar linha do header (com EDP, ...)
    header_row = None
    for i, row in enumerate(rows):
        if row and 'EDP' in [str(c).upper() if c else '' for c in row]:
            header_row = i
            break
    if header_row is None:
        return
    header = rows[header_row]
    # Mapear nome → coluna
    col_to_rubrica = {}
    for col_idx, name in enumerate(header):
        if not name:
            continue
        n = str(name).strip().upper()
        if 'EDP' in n:           col_to_rubrica[col_idx] = 'rub_edp'
        elif 'TELEFONE' in n:    col_to_rubrica[col_idx] = 'rub_telefone'
        elif 'SCHINDLER' in n:   col_to_rubrica[col_idx] = 'rub_schindler'
        elif 'ÁGUA' in n or 'AGUA' in n: col_to_rubrica[col_idx] = 'rub_agua'
        elif 'LIMPEZA' in n:     col_to_rubrica[col_idx] = 'rub_limpeza'
        elif 'BANC' in n:        col_to_rubrica[col_idx] = 'rub_banc'
        elif 'ALLIANZ' in n:     col_to_rubrica[col_idx] = 'rub_allianz'
        elif 'INTERVENÇÕES' in n or 'INTERVENCOES' in n or 'OBRAS' in n: col_to_rubrica[col_idx] = 'rub_obras'
        elif 'PLANO PAGAMENTO' in n or 'PLANO PAG' in n: col_to_rubrica[col_idx] = 'rub_plano_schindler'
        elif 'OUTRAS' in n or 'OUTROS' in n: col_to_rubrica[col_idx] = 'rub_outras'

    # Linhas Janeiro a Dezembro (12 meses)
    for mes in range(1, 13):
        row_idx = header_row + mes
        if row_idx >= len(rows):
            break
        row = rows[row_idx]
        if not row:
            continue
        for col_idx, rubrica_id in col_to_rubrica.items():
            if col_idx < len(row):
                val = row[col_idx]
                add_despesa(year, mes, val, rubrica_id, f"{MES_NAMES[mes-1]} {year} (histórico)")

for year, sheet_name in [(2021,'Despesas 2021'),(2022,'Despesas 2022'),(2023,'Despesas 2023'),
                          (2024,'Despesas 2024'),(2025,'Despesas 2025'),(2026,'Despesas 2026')]:
    parse_despesas(sheet_name, year)

print(f"Despesas geradas: {len(pagamentosDespesa)}")

# ──────────────────────────────────────────────────────────
# PLANOS · lidos EXATAMENTE das tabelas Excel
#
# Cada célula com valor numérico = 1 prestação registada como paga.
# A diferença entre quota total e total registado = valor em falta.
# Não criamos prestações "futuras" ou "pendentes" artificiais.
# ──────────────────────────────────────────────────────────

# Mapeamento de fração letra → tenantId (lendo Reparação Elevador 2025)
FRAC_LETRA_TO_TENANT = {
    'A': 'cond_02', 'B': 'cond_01', 'C': 'cond_04', 'D': 'cond_03',
    'E': 'cond_06', 'F': 'cond_05', 'G': 'cond_08', 'H': 'cond_07',
    'I': 'cond_10', 'J': 'cond_09',
}

planos = []
prestacoes = []

def parse_plano_sheet(sheet_name, plano_id, plano_nome, descricao,
                       linha_dados_inicio, col_dados_inicio,
                       col_frac_letra, col_quota_total, col_quota_mensal,
                       header_meses_row, datas_explicitas=None):
    """
    Lê uma folha de plano e gera:
     - 1 doc no array `planos` com campos exigidos pela UI
     - prestações em MATRIZ (cada tenant tem N prestações iguais ao nº de meses do plano)
       · estado='paga' se a célula tem valor, 'pendente' caso contrário
    """
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    if datas_explicitas:
        datas_meses = datas_explicitas
    else:
        header = rows[header_meses_row]
        datas_meses = []
        for c in range(col_dados_inicio, len(header)):
            val = header[c]
            if isinstance(val, datetime):
                datas_meses.append((c, val))

    num_prestacoes = len(datas_meses)
    inicio_data = min(d for _, d in datas_meses) if datas_meses else None
    fim_data = max(d for _, d in datas_meses) if datas_meses else None

    # Calcular valor total como soma das quotas totais das frações
    total_centimos_plano = 0
    quotas_por_tenant = {}  # tenant_id → quota_total_cent, quota_mensal_cent

    for r_idx in range(linha_dados_inicio, min(len(rows), linha_dados_inicio + 12)):
        row = rows[r_idx]
        if not row: continue
        letra = row[col_frac_letra] if col_frac_letra < len(row) else None
        if letra is None or letra == 'TOTAL': continue
        letra = str(letra).strip()
        if letra not in FRAC_LETRA_TO_TENANT: continue
        tenant_id = FRAC_LETRA_TO_TENANT[letra]
        try:
            qtotal = int(round(float(row[col_quota_total]) * 100))
        except (TypeError, ValueError):
            qtotal = 0
        try:
            qmensal = int(round(float(row[col_quota_mensal]) * 100))
        except (TypeError, ValueError):
            qmensal = qtotal // num_prestacoes if num_prestacoes else qtotal
        quotas_por_tenant[tenant_id] = (qtotal, qmensal, row)
        total_centimos_plano += qtotal

    # Gerar prestações em matriz · cada tenant tem N prestações
    total_pago_geral = 0
    for tenant_id, (qtotal, qmensal, row) in quotas_por_tenant.items():
        for i, (col_idx, data_mes) in enumerate(datas_meses):
            cell_val = row[col_idx] if col_idx < len(row) else None
            mes_ref = data_mes.strftime('%Y-%m')
            try:
                cell_eur = float(cell_val) if cell_val is not None and cell_val != '---' else None
            except (TypeError, ValueError):
                cell_eur = None

            if cell_eur and cell_eur > 0:
                # Pagamento real registado no Excel
                cell_cent = int(round(cell_eur * 100))
                prestacoes.append({
                    'id': f'prest_{plano_id}_{tenant_id}_{i+1}',
                    'planoId': plano_id,
                    'tenantId': tenant_id,
                    'numeroPrestacao': i + 1,
                    'mesReferencia': mes_ref,
                    'valor_centimos': cell_cent,
                    'estado': 'paga',
                    'pagoEm': data_mes.timestamp() * 1000,
                    'historico': True,
                })
                total_pago_geral += cell_cent
            else:
                # Sem valor · prestação pendente com valor mensal teórico
                prestacoes.append({
                    'id': f'prest_{plano_id}_{tenant_id}_{i+1}',
                    'planoId': plano_id,
                    'tenantId': tenant_id,
                    'numeroPrestacao': i + 1,
                    'mesReferencia': mes_ref,
                    'valor_centimos': qmensal,
                    'estado': 'pendente',
                    'historico': True,
                })

    estado_plano = 'concluido' if total_pago_geral >= total_centimos_plano * 0.999 else 'ativo'

    planos.append({
        'id': plano_id,
        'nome': plano_nome,
        'descricao': descricao,
        'valorTotal_centimos': total_centimos_plano,
        'numeroPrestacoes': num_prestacoes,
        'baseCalculo': 'manual',  # valores foram alocados manualmente por permilagem do Excel
        'dataInicio': inicio_data.strftime('%Y-%m') if inicio_data else None,
        'dataPrevisaoFim': fim_data.strftime('%Y-%m') if fim_data else None,
        'estado': estado_plano,
        'criadoEm': inicio_data.timestamp() * 1000 if inicio_data else None,
        'historico': True,
    })

# Reparação Elevador 2023 (Ago-Out 2023, 3 meses, headers em string)
parse_plano_sheet(
    sheet_name='Reparação Elevador 2023',
    plano_id='plano_elev_2023',
    plano_nome='Reparação Elevador 2023',
    descricao='Reparação extraordinária do elevador · 3 prestações (Ago-Out 2023)',
    linha_dados_inicio=3, col_dados_inicio=4,
    col_frac_letra=1, col_quota_total=7, col_quota_mensal=4,
    header_meses_row=2,
    datas_explicitas=[
        (4, datetime(2023, 8, 1)),
        (5, datetime(2023, 9, 1)),
        (6, datetime(2023, 10, 1)),
    ],
)

# Quotização Obras 2024 (Mar 2024 - Fev 2025, 12 prestações)
parse_plano_sheet(
    sheet_name='Quotização Obras',
    plano_id='plano_obras_2024',
    plano_nome='Quotização Obras 2024',
    descricao='Quotização extraordinária para obras de intervenção · 12 prestações (Mar 2024 - Fev 2025)',
    linha_dados_inicio=4, col_dados_inicio=6,
    col_frac_letra=1, col_quota_total=4, col_quota_mensal=5,
    header_meses_row=3,
    datas_explicitas=[
        (6,  datetime(2024, 3, 24)),  (7,  datetime(2024, 4, 24)),
        (8,  datetime(2024, 5, 24)),  (9,  datetime(2024, 6, 24)),
        (10, datetime(2024, 7, 24)),  (11, datetime(2024, 8, 24)),
        (12, datetime(2024, 9, 24)),  (13, datetime(2024, 10, 24)),
        (14, datetime(2024, 11, 24)), (15, datetime(2024, 12, 24)),
        (16, datetime(2025, 1, 24)),  (17, datetime(2025, 2, 24)),
    ],
)

# Reparação Elevador 2025 (12 colunas no header: Abr 2025 - Mar 2026)
parse_plano_sheet(
    sheet_name='Reparação Elevador 2025',
    plano_id='plano_elev_2025',
    plano_nome='Reparação Elevador 2025',
    descricao='Reparação extraordinária do elevador 2025',
    linha_dados_inicio=4, col_dados_inicio=6,
    col_frac_letra=1, col_quota_total=4, col_quota_mensal=5,
    header_meses_row=3,
)

print(f"Planos gerados: {len(planos)}")
for p in planos:
    pagos = sum(x['valor_centimos'] for x in prestacoes if x['planoId']==p['id'] and x['estado']=='paga')
    print(f"  • {p['nome']}: {p['valorTotal_centimos']/100:.2f}€ total · {pagos/100:.2f}€ pago · estado={p['estado']}")
print(f"Prestações geradas: {len(prestacoes)}")

print(f"Planos gerados: {len(planos)}")
print(f"Prestações geradas: {len(prestacoes)}")

# ──────────────────────────────────────────────────────────
# ORÇAMENTO 2026 APROVADO
# ──────────────────────────────────────────────────────────
orcamento_2026 = {
    'id': 'orc_2026_v1',
    'ano': 2026,
    'versao': 1,
    'estado': 'aprovado',
    'criadoEm': datetime(2025, 12, 1).timestamp() * 1000,
    'aprovadoEm': datetime(2025, 12, 15).timestamp() * 1000,
    'aprovadoPor': 'Ricardo Cordeiro',
    'arredondamento_centimos': 100,
    'incrementoPercentual': 0,
    'quotas': [
        {'tenantId': 'cond_02', 'permilage': 91, 'valorMensalAnterior': 3700, 'valorMensalNovo': 3700},
        {'tenantId': 'cond_01', 'permilage': 79, 'valorMensalAnterior': 3200, 'valorMensalNovo': 3200},
        {'tenantId': 'cond_04', 'permilage': 87, 'valorMensalAnterior': 3600, 'valorMensalNovo': 3600},
        {'tenantId': 'cond_03', 'permilage': 119,'valorMensalAnterior': 4800, 'valorMensalNovo': 4800},
        {'tenantId': 'cond_06', 'permilage': 88, 'valorMensalAnterior': 3600, 'valorMensalNovo': 3600},
        {'tenantId': 'cond_05', 'permilage': 121,'valorMensalAnterior': 4900, 'valorMensalNovo': 4900},
        {'tenantId': 'cond_08', 'permilage': 88, 'valorMensalAnterior': 3600, 'valorMensalNovo': 3600},
        {'tenantId': 'cond_07', 'permilage': 115,'valorMensalAnterior': 4700, 'valorMensalNovo': 4700},
        {'tenantId': 'cond_10', 'permilage': 87, 'valorMensalAnterior': 3600, 'valorMensalNovo': 3600},
        {'tenantId': 'cond_09', 'permilage': 125,'valorMensalAnterior': 5100, 'valorMensalNovo': 5100},
    ],
    'despesas': [
        {'rubricaId': 'rub_edp',          'orcado_centimos': 85352,  'realizado_centimos_anoAnterior': 77593},
        {'rubricaId': 'rub_agua',         'orcado_centimos': 23416,  'realizado_centimos_anoAnterior': 21287},
        {'rubricaId': 'rub_schindler',    'orcado_centimos': 140968, 'realizado_centimos_anoAnterior': 128153},
        {'rubricaId': 'rub_allianz',      'orcado_centimos': 84031,  'realizado_centimos_anoAnterior': 76392},
        {'rubricaId': 'rub_banc',         'orcado_centimos': 9972,   'realizado_centimos_anoAnterior': 9972},
        {'rubricaId': 'rub_limpeza',      'orcado_centimos': 120000, 'realizado_centimos_anoAnterior': 116400},
        {'rubricaId': 'rub_outras',       'orcado_centimos': 10000,  'realizado_centimos_anoAnterior': 0},
        {'rubricaId': 'rub_obras',        'orcado_centimos': 324835, 'realizado_centimos_anoAnterior': 0},
        {'rubricaId': 'rub_plano_schindler', 'orcado_centimos': 602326, 'realizado_centimos_anoAnterior': 0},
    ],
}

# ──────────────────────────────────────────────────────────
# META + Saldos
#
# Estrutura: meta vai conter docs com id (será expandido em coleção).
#   - condominio  · dados da entidade
#   - config      · saldoInicial por ano + saldoConhecido (ancoragem BPI)
# ──────────────────────────────────────────────────────────
meta = {
    'condominio': {
        'nome': 'Administração Condomínio Av. Amália Rodrigues, 24',
        'morada': 'Av. Amália Rodrigues, 24',
        'codigoPostal': '2650-437',
        'localidade': 'Amadora',
        'nif': '901589381',
        'iban': 'PT50 0010 0000 5398 9510 0019 7',
        'banco': 'BPI',
        'email': '',
        'telefone': '',
    },
    'config': {
        # Operadores admin · escolhidos no ecrã de login
        'administracao': {
            'nomes': ['Ricardo Nabais Cordeiro', 'Filipe Solha'],
            'emailContaCondominio': 'condoamira24@gmail.com',
        },
        # Saldo inicial de cada ano (em cêntimos) · base do cálculo
        'saldoInicial': {
            '2026': 321478,    # CO 2510.44 + Poup 704.34 (01-Jan-2026)
        },
        # Saldo real observado · ancoragem para detectar descalibração
        'saldoConhecido': {
            'data': '2026-05-25',
            'contaOrdem_centimos': 752178,    # 7521.78 €
            'contaPoupanca_centimos': 70434,  # 704.34 €
            'total_centimos': 822612,         # 8226.12 €
            'notas': 'BPI Net Empresas · posição integrada · ponto de ancoragem inicial',
            'registadoEm': int(datetime.now().timestamp() * 1000),
        },
        # Numeração de recibos · usada para gerar nº sequencial de novos recibos
        'nextNumberByYear': {
            '2026': 200,  # depois do último importado · começa em 200 para sair fora do range histórico
        },
    },
    # Plano Schindler · pagamentos faseados do condomínio à Schindler
    # Excel: Entrada 10% + 12 prestações mensais (Dez 2025 - Nov 2026) = 6023.26 €
    'planoSchindler': {
        'inicio': '2025-11-01',
        'fim': '2026-11-30',
        'fornecedor': 'Schindler',
        'descricao': 'Plano de pagamento faseado para reparação do elevador',
        'totalPrevisto_centimos': 602326,
        'rubricaId': 'rub_plano_schindler',
        'prestacoes': [
            {'data': '2025-11-01', 'valor_centimos': 62326, 'descricao': 'Entrada (10%)'},
            {'data': '2025-12-01', 'valor_centimos': 45000, 'descricao': 'Prestação 1 / 12 · Dez 2025'},
            {'data': '2026-01-01', 'valor_centimos': 45000, 'descricao': 'Prestação 2 / 12 · Jan 2026'},
            {'data': '2026-02-01', 'valor_centimos': 45000, 'descricao': 'Prestação 3 / 12 · Fev 2026'},
            {'data': '2026-03-01', 'valor_centimos': 45000, 'descricao': 'Prestação 4 / 12 · Mar 2026'},
            {'data': '2026-04-01', 'valor_centimos': 45000, 'descricao': 'Prestação 5 / 12 · Abr 2026'},
            {'data': '2026-05-01', 'valor_centimos': 45000, 'descricao': 'Prestação 6 / 12 · Mai 2026'},
            {'data': '2026-06-01', 'valor_centimos': 45000, 'descricao': 'Prestação 7 / 12 · Jun 2026'},
            {'data': '2026-07-01', 'valor_centimos': 45000, 'descricao': 'Prestação 8 / 12 · Jul 2026'},
            {'data': '2026-08-01', 'valor_centimos': 45000, 'descricao': 'Prestação 9 / 12 · Ago 2026'},
            {'data': '2026-09-01', 'valor_centimos': 45000, 'descricao': 'Prestação 10 / 12 · Set 2026'},
            {'data': '2026-10-01', 'valor_centimos': 45000, 'descricao': 'Prestação 11 / 12 · Out 2026'},
            {'data': '2026-11-01', 'valor_centimos': 45000, 'descricao': 'Prestação 12 / 12 · Nov 2026'},
        ],
    },
}

# ──────────────────────────────────────────────────────────
# OUTROS RECEBIMENTOS (Devolução Reabilita+ e Upgrade Videoporteiro 2026)
# ──────────────────────────────────────────────────────────
outrosRecebimentos = [
    {
        'id': 'or_reabilita_2026',
        'data': '2026-04-15',
        'descricao': 'Devolução Reabilita+',
        'valor_centimos': 651900,
        'metodoPagamento': 'transferencia',
        'origem': 'Reabilita+',
        'criadoEm': datetime(2026, 4, 15).timestamp() * 1000,
        'historico': True,
    },
    {
        'id': 'or_videoporteiro_2026',
        'data': '2026-03-20',
        'descricao': 'Upgrade Videoporteiro · contribuições dos condóminos',
        'valor_centimos': 14500,
        'metodoPagamento': 'transferencia',
        'origem': 'Condóminos',
        'criadoEm': datetime(2026, 3, 20).timestamp() * 1000,
        'historico': True,
    },
]

# ──────────────────────────────────────────────────────────
# Output
# ──────────────────────────────────────────────────────────

snapshot = {
    'meta': meta,
    'tenants': TENANTS,
    'rubricas': RUBRICAS,
    'receipts': receipts,
    'pagamentosDespesa': pagamentosDespesa,
    'planos': planos,
    'prestacoes': prestacoes,
    'orcamentos': [orcamento_2026],
    'outrosRecebimentos': outrosRecebimentos,
    'comunicacoes': [],
    '__importInfo': {
        'geradoEm': datetime.now().isoformat(),
        'origem': 'Backup Excel Condomínio AR24 v3.0',
        'periodo': '2021-2026',
        'contagens': {
            'tenants': len(TENANTS),
            'rubricas': len(RUBRICAS),
            'receipts': len(receipts),
            'despesas': len(pagamentosDespesa),
            'planos': len(planos),
            'prestacoes': len(prestacoes),
            'outrosRecebimentos': len(outrosRecebimentos),
        }
    }
}

with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(snapshot, f, ensure_ascii=False, indent=2)

print(f"\n✅ Snapshot gerado em: {OUT}")
print(f"Tamanho: {os.path.getsize(OUT) / 1024:.1f} KB")
print()
print("Sumário:")
for k, v in snapshot['__importInfo']['contagens'].items():
    print(f"  • {k}: {v}")
