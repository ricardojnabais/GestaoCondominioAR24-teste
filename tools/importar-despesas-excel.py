"""
Replica EXATA das tabelas Despesas 2024, 2025, 2026 do Excel principal.
Cada célula com valor > 0 vira uma entrada em pagamentosDespesa.
Apaga todas as despesas anteriores.
"""
import openpyxl, json
from datetime import datetime

EXCEL = "/mnt/user-data/uploads/Contas_Condominio_2023_2026_-_actualizado_a_25-05-2026.xlsx"
SEED = "/home/claude/GestaoCondominioAR24/data/seed-historico.json"

# Mapeamento POR NOME do header (case-insensitive) → rubricaId
NOME_TO_RUBRICA = {
    'edp': 'rub_edp',
    'schindler': 'rub_schindler',
    'água': 'rub_agua',
    'limpeza': 'rub_limpeza',
    'desp.banc.': 'rub_banc',
    'allianz': 'rub_allianz',
    'outras': 'rub_outras',
    'outros': 'rub_outras',
    'plano pagamento': 'rub_plano_schindler',
    'intervenções condomínio': 'rub_obras',
}

MES_PT = ['Janeiro','Fevereiro','Março','Abril','Maio','Junho',
          'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']

wb = openpyxl.load_workbook(EXCEL, data_only=True)

# Coletar todas as despesas
todas_despesas = []
for ano in ['2024', '2025', '2026']:
    ws = wb[f'Despesas {ano}']
    # Ler header L4 dinamicamente
    header = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    col_to_rubrica = {}
    for i, h in enumerate(header):
        if h is None: continue
        h_norm = str(h).strip().lower()
        if h_norm in NOME_TO_RUBRICA:
            col_to_rubrica[i] = NOME_TO_RUBRICA[h_norm]
    print(f"{ano} · cols mapeadas: {col_to_rubrica}")

    # L5 a L16 = Jan a Dez
    for mes_num in range(1, 13):
        row = list(ws.iter_rows(min_row=mes_num+4, max_row=mes_num+4, values_only=True))[0]
        for col_idx, rubricaId in col_to_rubrica.items():
            if col_idx >= len(row): continue
            val = row[col_idx]
            if val and isinstance(val, (int, float)) and val > 0:
                todas_despesas.append({
                    'ano': ano, 'mes': mes_num, 'rubricaId': rubricaId,
                    'valor_cent': int(round(val * 100)),
                })

print(f"Despesas no Excel (2024-2026): {len(todas_despesas)}")

# Sumário por ano e rúbrica
from collections import defaultdict
por_ano_rubrica = defaultdict(lambda: defaultdict(int))
por_ano = defaultdict(int)
for d in todas_despesas:
    por_ano_rubrica[d['ano']][d['rubricaId']] += d['valor_cent']
    por_ano[d['ano']] += d['valor_cent']

print(f"\nTotal por ano:")
for ano in ['2024','2025','2026']:
    print(f"  {ano}: {por_ano[ano]/100:>9.2f} €")

print(f"\nDetalhe 2026:")
for r, v in por_ano_rubrica['2026'].items():
    print(f"  {r}: {v/100:.2f} €")

# Carregar JSON
with open(SEED) as f:
    data = json.load(f)
print(f"\nDespesas no JSON antes: {len(data['pagamentosDespesa'])}")

# Apagar TODAS as despesas anteriores e criar novas
data['pagamentosDespesa'] = []

contador = 0
for d in todas_despesas:
    contador += 1
    data['pagamentosDespesa'].append({
        'id': f"pd_{d['ano']}_{d['mes']:02d}_{d['rubricaId']}",
        'rubricaId': d['rubricaId'],
        'valor_centimos': d['valor_cent'],
        'data': f"{d['ano']}-{d['mes']:02d}-15",
        'descricao': f"{[r for r in data['rubricas'] if r['id']==d['rubricaId']][0]['nome']} · {MES_PT[d['mes']-1]} {d['ano']}",
        'metodoPagamento': 'transferencia',
        'fornecedor': '',
        'criadoEm': datetime(int(d['ano']), d['mes'], 15).timestamp() * 1000,
        'historico': True,
    })

print(f"Despesas geradas: {len(data['pagamentosDespesa'])}")

# Salvar
with open(SEED, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)
print("✓ Snapshot atualizado")
